# -*- coding: utf-8 -*-
"""
Tesoreria App - Carga de datos de Requerimiento a Carga Real
Ejecutar con: streamlit run tesoreria_app.py
"""

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import io

# ============================================================
# CONFIGURACIÓN DE MAPEOS
# ============================================================

MAPEOS = {
    "03_Cuentas por Cobrar": {
        "hoja_carga_real": "Carga Real Ingresos",
        "fila_inicio": 9,
        "fuente": "Cuentas por Cobrar",
        "campos": [
            ("A", "D", None, None),   # Cliente → D
            ("B", "C", None, None),   # Documento → C
            ("H", "G", None, None),   # Probabilidad → G
            ("D", "H", None, None),   # Fecha vencimiento → H
            ("J", "E", None, None),   # Comentario → Descripción
            ("E", "L", None, None),   # Monto → L
        ],
        "categoria": "Cobranzas de ventas",
        "categoria_col": "K"
    },
    
    "04_Cuentas por Pagar": {
        "hoja_carga_real": "Carga Real Egresos Oper.",
        "fila_inicio": 9,
        "fuente": None,
        "campos": [
            ("C", "B", None, None),   # Fecha emisión → B
            ("B", "C", None, None),   # Documento → C
            ("A", "D", None, None),   # Proveedor → D
            ("J", "F", None, None),   # Comentario → Descripción
            ("H", "H", None, None),   # Prioridad pago → Prioridad Pago
            ("D", "G", None, None),   # Fecha vencimiento → G
            ("E", "L", None, None),   # Monto → L
        ],
        "categoria": "Pagos a proveedores",
        "categoria_col": "K"
    },
    
    "05_Ingresos Esperados": {
        "hoja_carga_real": "Carga Real Ingresos",
        "fila_inicio": 30,
        "fuente": "Ingresos esperados",
        "campos": [
            ("A", "D", None, None),   # Cliente → D
            ("B", "E", None, None),   # Concepto → Descripción
            ("F", "G", None, None),   # Probabilidad → G
            ("C", "H", None, None),   # Fecha esperada → H
            ("H", "E", None, None),   # Comentario → Descripción
            ("D", "L", None, None),   # Monto → L
        ],
        "categoria": "Cobranzas de ventas",
        "categoria_col": "K"
    },
    
    "07_Planilla": {
        "hoja_carga_real": "Carga Real Planilla",
        "fila_inicio": 9,
        "fuente": None,
        "campos": [
            ("B", "C", None, None),   # Área → C
            ("D", "D", None, None),   # Concepto → D
            ("E", "B", None, None),   # Fecha estimada → B
            ("F", "J", None, None),   # Monto → J
            ("H", "F", None, None),   # Prioridad pago → F
            ("J", "E", None, None),   # Comentario → E
        ],
        "categoria": "Planilla y beneficios sociales",
        "categoria_col": "I"
    },
    
    "06_Egresos Recurrentes": {
        "hoja_carga_real": "Carga Real Egresos Adm.",
        "fila_inicio": 9,
        "fuente": None,
        "campos": [
            ("A", "B", None, None),   # Categoría gasto → B
            ("B", "C", None, None),   # Concepto → C
            ("C", "D", None, None),   # Frecuencia → D
            ("J", "E", None, None),   # Comentario → Tipo de Gasto
            ("D", "F", None, None),   # Fecha esperada → F
            ("G", "G", None, None),   # Prioridad pago → G
            ("H", "H", None, None),   # Postergable → H
            ("E", "K", None, None),   # Monto → K
        ],
        "categoria": "Gastos operativos y fijos",
        "categoria_col": "J"
    },
}


# ============================================================
# FUNCIONES DE PROCESAMIENTO
# ============================================================

def analizar_requerimiento(file):
    """Analiza el archivo de requerimiento y retorna los datos."""
    wb = load_workbook(file, data_only=True)
    
    resultados = {}
    
    for hoja_req, config in MAPEOS.items():
        if hoja_req not in wb.sheetnames:
            continue
        
        ws = wb[hoja_req]
        
        # Leer datos (desde fila 5)
        datos = []
        for row in range(5, ws.max_row + 1):
            row_data = {}
            has_data = False
            for col in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row=row, column=col)
                    if hasattr(cell, 'column_letter'):
                        val = cell.value
                        if val is not None:
                            has_data = True
                            row_data[cell.column_letter] = val
                except:
                    pass
            
            if has_data:
                datos.append(row_data)
        
        resultados[hoja_req] = {
            "datos": datos,
            "config": config
        }
    
    return resultados


def aplicar_mapeos(wb_carga_real, analisis):
    """Aplica los mapeos al archivo de carga real.
    
    Si varias hojas del requerimiento van a la misma hoja destino,
    se copian secuencialmente una después de la otra.
    """
    
    cambios_realizados = []
    
    # Track fila actual para cada hoja destino
    fila_actual_por_hoja = {}
    
    # Procesar en el orden que vienen del analisis
    for hoja_req, info in analisis.items():
        config = info["config"]
        hoja_cr = config["hoja_carga_real"]
        fila_inicio_base = config.get("fila_inicio", 9)
        fuente = config.get("fuente", None)
        
        if hoja_cr not in wb_carga_real.sheetnames:
            continue
        
        ws_cr = wb_carga_real[hoja_cr]
        
        # Determinar fila de inicio:
        # - Si ya procesamos algo para esta hoja destino, continuar donde terminó
        # - Si no, usar la fila base del mapeo
        if hoja_cr in fila_actual_por_hoja:
            fila_inicio = fila_actual_por_hoja[hoja_cr]
        else:
            fila_inicio = fila_inicio_base
        
        # Procesar cada fila de datos del requerimiento
        for idx, row_data in enumerate(info["datos"]):
            fila_destino = fila_inicio + idx
            
            # Establecer categoría
            categoria_col_letter = config["categoria_col"]
            ws_cr[f"{categoria_col_letter}{fila_destino}"] = config["categoria"]
            
            cambios_realizados.append({
                "hoja": hoja_req,
                "hoja_cr": hoja_cr,
                "celda": f"{categoria_col_letter}{fila_destino}",
                "valor": config["categoria"],
                "tipo": "categoria"
            })
            
            # Establecer Fuente si está definida
            if fuente:
                ws_cr[f"I{fila_destino}"] = fuente
                cambios_realizados.append({
                    "hoja": hoja_req,
                    "hoja_cr": hoja_cr,
                    "celda": f"I{fila_destino}",
                    "valor": fuente,
                    "tipo": "fuente"
                })
            
            # Procesar cada campo del mapeo
            for col_req, col_cr, _, _ in config["campos"]:
                valor = row_data.get(col_req)
                
                if valor is None:
                    continue
                
                # Copiar valor
                cell_ref = f"{col_cr}{fila_destino}"
                ws_cr[cell_ref] = valor
                
                cambios_realizados.append({
                    "hoja": hoja_req,
                    "hoja_cr": hoja_cr,
                    "celda": cell_ref,
                    "valor": valor,
                    "tipo": "dato"
                })
            
            # Actualizar fila actual para esta hoja destino
            fila_actual_por_hoja[hoja_cr] = fila_destino + 1
    
    return cambios_realizados


# ============================================================
# INTERFAZ STREAMLIT
# ============================================================

def main():
    st.set_page_config(
        page_title="Tesoreria - Carga de Datos",
        page_icon="📊",
        layout="wide"
    )
    
    st.title("📊 Tesoreria - Carga de Datos")
    st.markdown("---")
    
    # Sidebar para configuración
    st.sidebar.header("Configuración")
    
    st.sidebar.markdown("### Mapeos activos")
    for hoja in MAPEOS.keys():
        st.sidebar.write(f"• {hoja}")
    
    # Inicializar variables de session state
    if 'carga_real_procesado' not in st.session_state:
        st.session_state['carga_real_procesado'] = None
    if 'nombre_archivo' not in st.session_state:
        st.session_state['nombre_archivo'] = None
    
    # Upload de archivos
    col1, col2 = st.columns(2)
    
    with col1:
        st.header("1. Archivo Requerimiento")
        
        requerimiento_file = st.file_uploader(
            "Subir archivo Requerimiento",
            type=["xlsx", "xls"],
            help="Archivo de requerimiento de información de la empresa"
        )
    
    with col2:
        st.header("2. Archivo Carga Real")
        
        carga_real_file = st.file_uploader(
            "Subir archivo Carga Real",
            type=["xlsx", "xls"],
            help="Archivo de plantilla Carga Real"
        )
    
    st.markdown("---")
    
    # Análisis
    if requerimiento_file:
        st.header("3. Análisis de archivos")
        
        if st.button("Analizar archivos", key="analizar_btn"):
            with st.spinner("Analizando..."):
                analisis = analizar_requerimiento(requerimiento_file)
                
                st.success(f"Se encontraron {len(analisis)} hojas con datos")
                
                # Mostrar resumen
                for hoja, info in analisis.items():
                    with st.expander(f"📋 {hoja}"):
                        st.write(f"**Hoja destino:** {info['config']['hoja_carga_real']}")
                        st.write(f"**Registros encontrados:** {len(info['datos'])}")
                        st.write(f"**Categoría:** {info['config']['categoria']}")
                        
                        # Mostrar primeros registros
                        if info['datos']:
                            df = pd.DataFrame(info['datos'][:5])
                            st.dataframe(df)
    
    # Aplicar cambios
    st.markdown("---")
    st.header("4. Aplicar cambios")
    
    aplicar_disabled = not (requerimiento_file and carga_real_file)
    
    if st.button("Aplicar cambios y descargar", type="primary", disabled=aplicar_disabled):
        if not requerimiento_file:
            st.error("Sube el archivo de Requerimiento primero")
        elif not carga_real_file:
            st.error("Sube el archivo de Carga Real primero")
        else:
            with st.spinner("Procesando..."):
                try:
                    # Cargar archivos
                    wb_cr = load_workbook(carga_real_file)
                    
                    # Procesar
                    analisis = analizar_requerimiento(requerimiento_file)
                    cambios = aplicar_mapeos(wb_cr, analisis)
                    
                    # Guardar en memoria
                    output = io.BytesIO()
                    wb_cr.save(output)
                    output.seek(0)
                    
                    st.session_state['carga_real_procesado'] = output.getvalue()
                    st.session_state['nombre_archivo'] = f"Carga_Real_Procesado_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                    
                    st.success(f"¡Cambios aplicados! Se procesaron {len(cambios)} celdas.")
                    
                    # Mostrar resumen
                    with st.expander("Ver resumen de cambios"):
                        cambios_datos = [c for c in cambios if c['tipo'] == 'dato']
                        for cambio in cambios_datos[:15]:
                            st.write(f"• {cambio['hoja_cr']}!{cambio['celda']}: {cambio['valor']}")
                        if len(cambios_datos) > 15:
                            st.write(f"... y {len(cambios_datos) - 15} cambios más")
                    
                except Exception as e:
                    st.error(f"Error: {str(e)}")
    
    # Descargar
    if st.session_state['carga_real_procesado']:
        st.download_button(
            label="Descargar Carga Real procesada",
            data=st.session_state['carga_real_procesado'],
            file_name=st.session_state['nombre_archivo'],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


if __name__ == "__main__":
    main()
